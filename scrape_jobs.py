import asyncio
from pyppeteer import launch
from openpyxl import load_workbook
import json

async def scrape_job_description(url):
    executable_path = r'C:\Users\hp\Downloads\chrome-win'
    browser = await launch(executablePath=executable_path)
    page = await browser.newPage()
    await page.goto(url)
    # Scrape job descriptions (modify slector as needed)
    job_descriptions = await page.evaluate('''() => {
        return Array.from(document.querySelectorAll('.job-description p')).map(p => p.innerText);
    }''')
    await browser.close()
    return job_descriptions

async def main():
    wb = load_workbook(filename = 'D:\web_scrapper_career\Exel11.xlsx')
    sheet = wb.active
    urls = [row[3] for row in sheet.iter_rows(min_row=2, values_only=True) if row[3] is not None] #URLs are in the fourth column, skipping header

    jobs_data = {}
    for url in urls:
            print(f"Scraping {url}")
            descriptions = await scrape_job_description(str(url))
            jobs_data[url] = descriptions

    with open('jobs_data.json', 'w') as json_file:
        json.dump(jobs_data, json_file, indent=4)

if __name__ == '__main__':
    asyncio.run(main())